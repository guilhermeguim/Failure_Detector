from time import sleep
import snap7.client as c
from snap7.util import *
from snap7.types import *
import time
import xlsxwriter
from pathlib import Path
from openpyxl import load_workbook
#from gpiozero import CPUTemperature
import threading
from multiprocessing import Process, Manager
from threading import Thread

import pandas as pd
import sqlite3
from read import read_total, get_reading
from calculate_parameters import find_area, calcula_bit, to_list, get_curr_path, calcula_bit_wc
import os
from join import join_files
import sys

global running
running = 1


curr_path = get_curr_path()

for file in os.listdir(curr_path): # percorre os arquivos do diretório local
    if file.endswith(".db"): # verifica se o arquivo termina com .db
        db_name = file # salva o nome do arquivo na variável db_name
        break # interrompe o loop
    
db_name = curr_path + str(db_name)

print(db_name)

conn = sqlite3.connect(db_name)

dataset2 = pd.read_sql_query('SELECT * FROM GENERAL_PARAMETERS', conn)
parameters2 = dataset2.values.tolist()

IP = str(parameters2[0][13])
RACK = int(parameters2[0][14])
SLOT = int(parameters2[0][15])
TCPPORT = int(parameters2[0][16])

global wc_data
global input_data
global wc_data_alt
global manual
global alarm

wc_data = 0
wc_data_alt = 0
input_data = 0
manual = 0
alarm = 0


def register_time():
    file_path = curr_path + 'logs/training.log'
    with open(file_path, "w") as f:
        # Obtém a data e hora atual
        agora = datetime.now()
        # Formata a data e hora como uma string
        data_hora = agora.strftime("%Y/%m/%d %H:%M:%S.%f")
        # Escreve a data e hora na primeira linha do arquivo
        f.write(data_hora)

def clean_directory():
        # Define o caminho da pasta
        pasta = curr_path + 'datasets'

        # Lista todos os arquivos .xlsx na pasta
        arquivos = os.listdir(pasta)

        # Percorre cada arquivo e verifica se é .xlsx
        for arquivo in arquivos:
        # Se for .xlsx, o apaga
                if arquivo.endswith('.xlsx'):
                        os.remove(os.path.join(pasta, arquivo))
                        #print(f'Arquivo {arquivo} apagado.')

def scan_in(minutos,bit_to_del,byte_start,size_word,total_size, t_num, wc_type,wc_position,wc_pat,ext_byte,detect_filter, d,region,size_manual,manual_bits, size_alarm, alarm_bits):
        
        print('STARTED: ', t_num)
        save_path =  curr_path + 'datasets/' + t_num + '.xlsx'
        
        if wc_type != 'word':
                wc_position = wc_position + 1
        
        try:
                workbook_name = Path(save_path)
                wb = load_workbook(workbook_name)
                page = wb.active
                wb.save(filename=workbook_name)
                
        except:
                path_file = Path(save_path)
                workbook = xlsxwriter.Workbook(path_file)
                worksheet = workbook.add_worksheet()
                workbook.close()
                
                workbook_name = Path(save_path)
                wb = load_workbook(workbook_name)
                page = wb.active
                wb.save(filename=workbook_name)
        
        #Variaveis temporárias de programa
        list_hex = [] #lista dos hexadecimais lidos pelo programa, essa lista será adicionada em uma linha de excel assim que finalizar o ciclo do plc
        row = 0 #itera pelas linhas do arquivo excel
        list_one = [] #lista do work complete, é utilizada para que o programa não fique pulando linhas no arquivo excel
        list_falha = []
        contador = 0 #define a duração da coleta
        time_start = time.time()
        minutos = minutos
        time_duration = 60*minutos
        
        #bits 10.4 e 10.5s
        bit_to_del = bit_to_del
        size_word = size_word
        
        #cpu = CPUTemperature()
        
        input_area = 'input_data' + str(region)
        start_area_str = input_area + '_start'
        start_area = d[start_area_str]

        read_end = time.time()
        while(True):

                #print('thread: ', t_num, 'time: ',time.time())
                
                #Inicia o contador do tempo para verificar o scan time
                start_time = time.time()
                #print(t_num,"--- Waiting %s seconds ---" % (start_time - read_end))
                #sleep(0.35)
                #print(t_num,'start :',start_time)
                #Incrementa o contador da duração da coleta
                contador = contador+1
                
                
                hex_word, wc_bit, manual_bit, alarm_bit = get_reading(d,byte_start,size_word,ext_byte,bit_to_del,wc_type,wc_position,total_size,input_area,start_area,size_manual,manual_bits, size_alarm, alarm_bits)
                
                #Contador do tamanho das listas auxiliares
                size = int(len(list_hex))
                size_one = int(len(list_one))
                
                #Roda o programa até a duração definida na variavel 'contador'
                if  time.time() < time_start+time_duration:
                        #print(contador)
                        if wc_bit == wc_pat : #se o plc está rodando uma sequencia
                                if manual_bit == 0 and alarm_bit == 0:
                                        list_one = [] #zera a lista de work complete (evita pular linhas no excel)
                                        list_falha = []
                                        if size == 0: # e se a lista de leituras estiver vazia

                                                list_hex.append(hex_word) #adicionar leitura na lista
                                        else: #se não estiver vazia
                                                if list_hex[size-1] != hex_word: #se o valor lido for diferente do ultimo elemento da lista
                                                        list_hex.append(hex_word) #adicionar leitura na lista
                                else:
                                        if manual_bit == 1 and alarm_bit == 0:
                                                tipo = 'manual'
                                                now = datetime.now()
                                                tempo =  str(now.strftime("%d/%m/%Y %H:%M:%S"))
                                        elif manual_bit == 0 and alarm_bit == 1:
                                                tipo = 'alarme'
                                                now = datetime.now()
                                                tempo =  str(now.strftime("%d/%m/%Y %H:%M:%S"))
                                        else:
                                                tipo = 'falha'
                                                now = datetime.now()
                                                tempo =  str(now.strftime("%d/%m/%Y %H:%M:%S"))
                                        size_falha = int(len(list_falha))
                                        if size_falha == 0:
                                                #print("ENTREI AQUI")
                                                
                                                list_hex = [tipo,tempo]
                                                workbook_name = Path(save_path)
                                                wb = load_workbook(workbook_name)
                                                page = wb.active
                                                page.append(list_hex)
                                                #worksheet.write_row(row,0,list_hex) #armazenar lista de leituras em uma linha de excel
                                                row = row+1 #joga para a linha de baixo
                                                list_hex = [] #zera a lista de leituras
                                                list_falha.append(wc_bit)
                                                
                                                wb.save(filename=workbook_name)
                                        else:
                                                if list_falha[size_falha-1] == wc_bit: #se o valor lido for igual do ultimo elemento da lista
                                                        pass
                        else: #se a sequencia parar
                                size_one = int(len(list_one))
                                if size_one == 0:

                                        workbook_name = Path(save_path)
                                        wb = load_workbook(workbook_name)
                                        page = wb.active
                                        page.append(list_hex)
                                        #worksheet.write_row(row,0,list_hex) #armazenar lista de leituras em uma linha de excel
                                        row = row+1 #joga para a linha de baixo
                                        list_hex = [] #zera a lista de leituras
                                        list_one.append(wc_bit)

                                        wb.save(filename=workbook_name)
                                else:
                                        if list_one[size_one-1] == wc_bit: #se o valor lido for igual do ultimo elemento da lista
                                                pass


                else:
                        workbook_name = Path(save_path)
                        wb = load_workbook(workbook_name)
                        page = wb.active
                        page.append(list_hex)
                        ##IMPLEMENTAR APPEND AQUI
                        wb.save(filename=workbook_name)
                        break
                
                # print(list_hex)
                # print(list_one)
                # print(size)
                # print(t_num,'end :',time.time())
                read_end = time.time()
                #if t_num == 400.0:
                #print(t_num,"Scan: --- %s seconds ---" % (read_end - start_time))
                # print("--- %s seconds ---" % (time.time() - time_start))
                # print("Read:", row-1)
                # print("--------------------------------------")

def scan_cicle(tempo, db_name, name_list):
        
        
        manager = Manager()
        shared_dict = manager.dict()
        
        conn = sqlite3.connect(db_name)
        
        dataset1 = pd.read_sql_query('SELECT * FROM GENERAL_PARAMETERS', conn)
        parameters1 = dataset1.values.tolist()
        
        print(parameters1)

        wc_address = int(parameters1[0][1])
        wc_size = int(parameters1[0][2])
        try:
                wc_alternative_address = int(parameters1[0][3])
                wc_alternative_size = int(parameters1[0][4])
        except:
                wc_alternative_address = parameters1[0][3]
                wc_alternative_size = parameters1[0][4]
        manual_types = to_list(parameters1[0][5])
        manual_addresses = to_list(parameters1[0][6])
        manual_bits = to_list(parameters1[0][7])
        alarm_types = to_list(parameters1[0][8])
        alarm_addresses = to_list(parameters1[0][9])
        alarm_bits = to_list(parameters1[0][10])
        
        start_list =  to_list(parameters1[0][11])
        total_size_list = to_list(parameters1[0][12])
        
        size_manual = len(manual_addresses)
        size_alarm = len(alarm_addresses)
        
        t_main = Process(target=read_total, 
                                args=(IP, RACK, SLOT, TCPPORT, 
                                shared_dict,
                                wc_address,
                                wc_size,
                                wc_alternative_address,
                                wc_alternative_size,
                                manual_types,
                                manual_addresses,
                                alarm_types,
                                alarm_addresses,
                                start_list, 
                                total_size_list))
        
        
        
        names = name_list
        names_str = "','".join(names)
        query = "SELECT * FROM PARAMETERS WHERE ANALISYS_NAME IN ('{}');".format(names_str)
        
        dataset = pd.read_sql_query(query, conn)
        parameters = dataset.values.tolist()
        
        process_list = []
        
        t_main.start()
        sleep(3)
        
        t_num_list = []
        
        for parameter in parameters:
                print(parameter)
                t = parameter[4]
                tempo = tempo
                start_byte = parameter[2]
                size_word = parameter[3]
                silent_bit = calcula_bit(parameter[1], start_byte, size_word)
                region, total_size = find_area(start_list, total_size_list,start_byte)
                thread_name = parameter[4]
                wc_type = parameter[5]
                wc_position = parameter[6]
                wc_pat = parameter[7]
                wc_ext_byte = parameter[8]
                detect_filter = parameter[9]
                
                t_num_list.append(thread_name)
                
                if wc_type == 'input':
                        wc_position_calculated = wc_position
                elif wc_type == 'memory':
                        wc_position_calculated = calcula_bit_wc(wc_address, wc_ext_byte, wc_position)
                elif wc_type == 'memory_alternative':
                        wc_position_calculated = calcula_bit_wc(wc_alternative_address, wc_ext_byte, wc_position)
                        
                
                thread = Process(target=scan_in, 
                args = (tempo,
                        silent_bit,
                        start_byte,
                        size_word,
                        total_size,
                        thread_name,
                        wc_type,
                        wc_position_calculated,
                        wc_pat,
                        wc_ext_byte,
                        detect_filter,
                        shared_dict,
                        region,
                        size_manual ,manual_bits,
                        size_alarm, alarm_bits))
                
                thread.start()
                process_list.append(thread)
                
        for p in process_list:
                p.join()
                
        return t_num_list

def delete_first_and_last_rows(filepath, path_training):
        # Ler arquivo excel
        df = pd.read_excel(filepath, header=None)
        
        # Verificar se DataFrame está vazio
        if df.empty:
                os.remove(filepath)
                return df

        # Adicionar linha vazia no início do DataFrame
        df = pd.concat([pd.DataFrame(['']*df.shape[1], df.columns).T, df], ignore_index=True)

        # Selecionar linhas que contém as palavras 'falha', 'manual' ou 'alarme'
        mask = df.iloc[:, 0].isin(['falha', 'manual', 'alarme'])

        # Selecionar linhas acima e abaixo das linhas selecionadas
        mask_shift = mask.shift(periods=-1) | mask.shift(periods=1) | mask

        # Excluir linhas selecionadas
        df = df[~mask_shift]
        
        # Excluir primeira e última linhas
        df = df.iloc[1:-1]

        if df.empty:
                os.remove(filepath)
                return df

        # Verificar se arquivo de treinamento existe
        if os.path.exists(path_training):
                training_df = pd.read_excel(path_training, header=None)
        else:
                training_df = pd.DataFrame()
                
        # Concatenar o DataFrame atualizado com o DataFrame do arquivo de treinamento
        training_df = pd.concat([training_df, df], ignore_index=True, axis=0)

        # Salvar arquivo de treinamento
        training_df.to_excel(path_training, index=False, header=False)
        
        os.remove(filepath)

def get_path():
        db_name = None # inicializa a variável db_name como None

        print(os.listdir(curr_path))

        for file in os.listdir(curr_path): # percorre os arquivos do diretório local
                if file.endswith(".db"): # verifica se o arquivo termina com .db
                        db_name = file # salva o nome do arquivo na variável db_name
                        break # interrompe o loop
                
        db_path = curr_path + str(db_name)

        print(db_path)
        

        log_path = curr_path + 'logs/alarm.log'
        
        return db_path, log_path

def check_empty():
        # Define o diretório
        directory = curr_path + 'datasets/training'

        # Lista todos os arquivos no diretório
        files = os.listdir(directory)

        # Cria uma variável para armazenar o resultado
        result = 1

        # Percorre cada arquivo e verifica se é .xlsx
        for file in files:
                # Se for .xlsx, altera o resultado para 2 e sai do loop
                if file.endswith('.xlsx'):
                        result = 2
                        break

        # Retorna o resultado
        return result

def keep_alive():
    while True:
        global running
        if running == 1:
                sleep(1)
                register_time()
        else:
                break

if __name__=="__main__":
        
        t_alive = Thread(target=keep_alive)
        t_alive.start()
        
        clean_directory()

        db_name, log_path = get_path()
        conn = sqlite3.connect(db_name)
        
        tempo = int(sys.argv[1])
        
        c = conn.cursor()

        # Execute a SELECT statement to retrieve all values from a column of a table
        c.execute('SELECT ANALISYS_NAME FROM PARAMETERS')

        # Fetch all rows and store the values in a list
        column_values = [row[0] for row in c.fetchall()]

        # Close the cursor and connection objects
        c.close()
        conn.close()

        # Print the list of column values
        print(column_values)

        size_columns = len(column_values)
        it = 0
        
        t_num_list = scan_cicle(tempo,db_name,column_values[it : size_columns])
        
        #opc=1 -> Primeiro treinamento, não existe nenhum arquivo em '/training
        #opc=2 -> Apenas adição de novos treinamentos
        opc = check_empty()
        join_files(opc)
        
        clean_directory()
        print('Ended')
        
        running = 0
        
        exit()


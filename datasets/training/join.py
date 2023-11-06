import pandas as pd
import numpy as np
from os import listdir
from os.path import isfile, join
from openpyxl import load_workbook, Workbook

def join_files(opc):
    
    mypath1 = 'C:/Users/37107014/Desktop/Q30_Engine_Failure_Detector-Validation_PC/datasets/training'
    mypath2 = 'C:/Users/37107014/Desktop/Q30_Engine_Failure_Detector-Validation_PC/datasets'
    
    if opc == 1:
        onlyfiles = [f for f in listdir(mypath1) if isfile(join(mypath1, f))]
        for file in onlyfiles:
            if file[-1] != 'y':
                try:
                    path1 = 'C:/Users/37107014/Desktop/Q30_Engine_Failure_Detector-Validation_PC/datasets/training/' + file
                    print(path1)
                    wb2 = load_workbook(path1)
                    ws2 = wb2['Sheet1']
                    nrows = ws2.max_row
                    ws2.delete_rows(nrows)
                    ws2.delete_rows(1)   
                    wb2.save(path1)
                except:
                    path1 = 'C:/Users/37107014/Desktop/Q30_Engine_Failure_Detector-Validation_PC/datasets/training/' + file
                    print('ERRO: ', path1)
                    
        
    print("##############################")
    
    onlyfiles = [f for f in listdir(mypath2) if isfile(join(mypath2, f))]
    for file in onlyfiles:
        if file[-1] != 'y':
            try:
                path2 = 'C:/Users/37107014/Desktop/Q30_Engine_Failure_Detector-Validation_PC/datasets/' + file
                print(path2)
                wb2 = load_workbook(path2)
                ws2 = wb2['Sheet1']
                nrows = ws2.max_row
                ws2.delete_rows(nrows)
                ws2.delete_rows(1)   
                wb2.save(path2)
            except:
                path2 = 'C:/Users/37107014/Desktop/Q30_Engine_Failure_Detector-Validation_PC/datasets/' + file
                print('ERRO: ',path2)
            
            
    print('acabei aqui')
            
    onlyfiles = [f for f in listdir(mypath1) if isfile(join(mypath1, f))]
    for file in onlyfiles:
        if file[-1] != 'y':
            try:
                path1 = 'C:/Users/37107014/Desktop/Q30_Engine_Failure_Detector-Validation_PC/datasets/training/' + file
                path2 = 'C:/Users/37107014/Desktop/Q30_Engine_Failure_Detector-Validation_PC/datasets/' + file
                print('concat: ',path1)
                # filenames
                excel_names = [path1, path2]
                # read them in
                excels = [pd.ExcelFile(name) for name in excel_names]
                # turn them into dataframes
                frames = [x.parse(x.sheet_names[0], header=None,index_col=None) for x in excels]
                # concatenate them..
                combined = pd.concat(frames)
                # write it out
                combined.to_excel(path1, header=False, index=False)
            except:
                print('CONCAT ERROR: ', file)
    print('concatenated')
    
    onlyfiles = [f for f in listdir(mypath1) if isfile(join(mypath1, f))]
    for file in onlyfiles:
        if file[-1] != 'y':
            try:
                path1 = 'C:/Users/37107014/Desktop/Q30_Engine_Failure_Detector-Validation_PC/datasets/training/' + file
                df = pd.read_excel(path1, header=None)
                
                col_list = df.columns.to_list()
        
                to_del = []
                
                to_del_falha = df.index[df[col_list[0]] == 'falha'].tolist()
                to_del_manual = df.index[df[col_list[0]] == 'manual'].tolist()
                to_del_alarme = df.index[df[col_list[0]] == 'alarme'].tolist()
                
                for i in to_del_falha:
                    to_del.append(i)
                
                for i in to_del_manual:
                    to_del.append(i)
                    
                for i in to_del_alarme:
                    to_del.append(i)
                
                to_add = []
                
                for i in to_del:
                    to_add.append(i+1)
                    
                for i in to_add:
                    to_del.append(i)
                    
                to_del = sorted(to_del, reverse=True)
                
                print(file)
                print(to_del)
                
                try:
                    df = df.drop(index=to_del)
                except:
                    df = df.drop(index=to_del[1:])
                    
                df.to_excel(path1, header=False, index=False)
            except:
                print('CLEAN ERROR: ', file)
                
                
if __name__ == '__main__':
    
    print('Option 1: Clean 1 and Join 2')
    print('Option 2: Only Join 2')
    
    option = input('Type option: ')
    join_files(int(option))
    
    